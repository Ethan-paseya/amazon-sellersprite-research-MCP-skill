"""Build the sanitized data-driven Product-Planning V1 workbook template."""

from __future__ import annotations

import argparse
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

import render_product_planning_v1 as renderer


def placeholder_distribution(labels: list[str]) -> dict:
    return {
        "items": [
            {"label": label, "products": 0, "share": 0.0, "evidenceId": f"DIST-{index:02d}"}
            for index, label in enumerate(labels, 1)
        ]
    }


def build_placeholder_evidence() -> dict:
    competitors = []
    for index in range(1, 5):
        competitors.append(
            {
                "asin": f"COMPETITOR_ASIN_{index}",
                "source": "user" if index <= 3 else "system",
                "brand": f"{{BRAND_{index}}}",
                "title": f"{{TITLE_{index}}}",
                "price": 0.0,
                "rating": 0.0,
                "ratings": 0,
                "reviews": 0,
                "fulfillment": f"{{FULFILLMENT_{index}}}",
                "listingDate": f"{{LISTING_DATE_{index}}}",
                "categoryPath": f"{{CATEGORY_PATH_{index}}}",
                "nodeIdPath": f"{{NODE_ID_PATH_{index}}}",
                "role": f"{{COMPETITOR_ROLE_{index}}}",
                "productAction": f"{{PRODUCT_MANAGER_ACTION_{index}}}",
                "features": f"{{VERIFIED_FEATURES_{index}}}",
                "positiveThemes": f"{{POSITIVE_REVIEW_THEMES_{index}}}",
                "negativeThemes": f"{{LOW_STAR_PAIN_POINTS_{index}}}",
                "priority": f"{{PRIORITY_{index}}}",
                "salesSnapshot": "N/A",
                "evidenceIds": [f"COMP-{index:02d}"],
                "imageUrl": "N/A",
            }
        )

    keywords = [
        {
            "keyword": f"{{KEYWORD_{index}}}",
            "searches": 0,
            "purchases": 0,
            "purchaseRate": 0.0,
            "ppc": 0.0,
            "products": 0,
            "supplyDemandRatio": 0.0,
            "tool": "keyword_miner",
            "evidenceId": f"KW-{index:03d}",
        }
        for index in range(1, 8)
    ]
    trend = [
        {
            "month": f"{{TREND_MONTH_{index:02d}}}",
            "avgUnits": 0,
            "avgRevenue": 0.0,
            "avgBsr": 0,
            "avgPrice": 0.0,
            "newProductShare": 0.0,
            "newAvgUnits": 0,
            "evidenceId": f"TREND-{index:02d}",
        }
        for index in range(1, 13)
    ]
    aba = []
    for index in range(1, 6):
        aba.append(
            {
                "keyword": f"{{ABA_KEYWORD_{index}}}",
                "202601": {"rank": 0, "searches": 0},
                "202602": {"rank": 0, "searches": 0},
                "202603": {"rank": 0, "searches": 0},
            }
        )

    data_sources = [
        ("SRC-ASIN", "asin_detail", "3 个用户种子 + 1 个系统补充竞品"),
        ("SRC-CANDIDATE", "competitor_lookup / product_research", "系统补充竞品候选池"),
        ("SRC-MARKET", "market_research_statistics", "目标叶子节点市场快照与趋势"),
        ("SRC-KEYWORD", "keyword_miner / keyword_research", "目标产品关键词"),
        ("SRC-VOC", "review", "4-5 星卖点与 1-3 星痛点主题"),
        ("SRC-ABA-MANUAL", "manual_input", "ABA 排名与搜索量由用户人工填入；自动流程不写入第五 Sheet"),
    ]
    return {
        "render": {
            "templateMode": True,
            "workbookTitle": "Product-Planning V1 数据驱动通用模板",
            "markdownTitle": "Product-Planning V1 数据驱动通用模板",
        },
        "command": {
            "site": "{SITE}",
            "seedAsins": ["SEED_ASIN_1", "SEED_ASIN_2", "SEED_ASIN_3"],
            "systemSupplement": "SYSTEM_SUPPLEMENT_ASIN",
        },
        "queryManifest": {
            "selectedNodeIdPath": "{NODE_ID_PATH}",
            "lastCompleteMonth": "{LAST_COMPLETE_MONTH}",
        },
        "productDirection": {
            "name": "{PRODUCT_DIRECTION}",
            "description": "{PRODUCT_DIRECTION_DESCRIPTION}",
            "targetPrice": "{TARGET_PRICE_BAND}",
            "targetAudience": "{TARGET_AUDIENCE}",
            "verifiedFunctions": [f"{{VERIFIED_FUNCTION_{index}}}" for index in range(1, 5)],
            "unmetNeeds": [f"{{UNMET_NEED_{index}}}" for index in range(1, 4)],
        },
        "categorySelection": {
            "selected": {
                "labelPath": "{CATEGORY_PATH}",
                "nodeIdPath": "{NODE_ID_PATH}",
                "reason": "运行时根据已验证竞品类目路径选择重复出现且与产品形态直接相关的最深叶子节点",
                "overrideReason": "如需改用父节点或相邻节点，必须记录覆盖率、相关性和拒绝其他节点的原因",
            }
        },
        "market": {
            "snapshot": {
                "month": "{SNAPSHOT_MONTH}",
                "products": 0,
                "brands": 0,
                "sellers": 0,
                "avgPrice": 0.0,
                "avgUnits": 0,
                "avgRevenue": 0.0,
                "avgBsr": 0,
                "avgRatings": 0,
                "avgRating": 0.0,
                "newProductShare": 0.0,
                "newAvgUnits": 0,
            },
            "monthlyTrend": trend,
            "demandTrend": {"handling": "运行时记录接口真实返回或空返回，不使用 ASIN 聚合替代类目趋势"},
            "concentration": {"brandCr5Units": 0.0},
            "distributions": {
                "listingAge": placeholder_distribution([f"{{LISTING_AGE_BUCKET_{index}}}" for index in range(1, 10)]),
                "listingYear": placeholder_distribution([f"{{LISTING_YEAR_BUCKET_{index}}}" for index in range(1, 11)]),
                "ratingCount": placeholder_distribution([f"{{RATING_COUNT_BUCKET_{index}}}" for index in range(1, 8)]),
                "rating": placeholder_distribution([f"{{RATING_BUCKET_{index}}}" for index in range(1, 8)]),
                "price": placeholder_distribution([f"{{PRICE_BUCKET_{index}}}" for index in range(1, 10)]),
            },
        },
        "competitors": competitors,
        "keywords": {"items": keywords, "missing": []},
        "abaQuarter": aba,
        "dataGaps": [
            {
                "field": "{DATA_GAP_FIELD_1}",
                "description": "{DATA_GAP_DESCRIPTION_1}",
                "impact": "{DATA_GAP_IMPACT_1}",
                "handling": "{DATA_GAP_HANDLING_1}",
            }
        ],
        "dataSources": [
            {
                "evidenceId": evidence_id,
                "tool": tool,
                "scope": scope,
                "status": "运行时填写",
                "note": "仅保留脱敏查询范围、采用状态和证据 ID",
            }
            for evidence_id, tool, scope in data_sources
        ],
        "decision": {
            "status": "{DECISION_STATUS}",
            "summary": "{DECISION_SUMMARY}",
            "gates": [],
            "actions": [
                {
                    "priority": f"P{index}",
                    "action": f"运行时生成第 {index + 1} 条可回查证据的产品经理动作",
                    "evidenceIds": [f"ACTION-{index + 1:02d}"],
                }
                for index in range(3)
            ],
        },
        "ipSignals": [],
    }


def clear_range(sheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            sheet.cell(row, column).value = None


def scrub_runtime_values(workbook) -> None:
    intent = workbook["意向产品"]
    intent._images = []
    intent_values = [
        "{FIRST_SEED_MAIN_IMAGE_RUNTIME}",
        "{TARGET_PRICE_BAND}",
        "{TARGET_AUDIENCE}",
        *[f"{{VERIFIED_FUNCTION_{index}}}" for index in range(1, 5)],
        *[f"{{UNMET_NEED_{index}}}" for index in range(1, 4)],
    ]
    for column, value in enumerate(intent_values, 1):
        intent.cell(5, column, value)

    market = workbook["市场分析"]
    for row in range(6, 18):
        market.cell(row, 1, f"{{TREND_MONTH_{row - 5:02d}}}")
        clear_range(market, row, row, 2, 4)
    for row in range(6, 13):
        market.cell(row, 13, f"{{KEYWORD_{row - 5}}}")
        clear_range(market, row, row, 14, 17)
    for start_row, count, start_col, prefix in (
        (24, 9, 1, "LISTING_AGE_BUCKET"),
        (24, 10, 13, "LISTING_YEAR_BUCKET"),
        (40, 7, 1, "RATING_COUNT_BUCKET"),
        (40, 7, 13, "RATING_BUCKET"),
        (53, 9, 1, "PRICE_BUCKET"),
    ):
        for offset in range(count):
            market.cell(start_row + offset, start_col, f"{{{prefix}_{offset + 1}}}")
            clear_range(market, start_row + offset, start_row + offset, start_col + 1, start_col + 2)

    competitors = workbook["竞品分析与优化策略"]
    for index, row in enumerate(range(6, 10), 1):
        values = [
            f"{{COMPETITOR_ASIN_{index}}}",
            "用户种子" if index <= 3 else "系统补充",
            f"{{BRAND_{index}}}",
            f"{{TITLE_{index}}}",
            None,
            None,
            None,
            f"{{FULFILLMENT_{index}}}",
            f"{{LISTING_DATE_{index}}}",
            f"{{CATEGORY_PATH_{index}}}",
            f"{{COMPETITOR_ROLE_{index}}}",
            f"{{PRODUCT_MANAGER_ACTION_{index}}}",
        ]
        for column, value in enumerate(values, 1):
            competitors.cell(row, column, value)
        competitors.cell(12 + index, 10, f"{{COMPETITOR_ASIN_{index}}}")
        competitors.cell(12 + index, 11, None)
        competitors.cell(12 + index, 12, None)
        voc_values = [
            f"{{COMPETITOR_ASIN_{index}}}",
            f"{{VERIFIED_FEATURES_{index}}}",
            f"{{POSITIVE_REVIEW_THEMES_{index}}}",
            f"{{LOW_STAR_PAIN_POINTS_{index}}}",
            f"{{PRODUCT_MANAGER_ACTION_{index}}}",
            f"{{PRIORITY_{index}}}",
            f"{{EVIDENCE_IDS_{index}}}",
        ]
        for column, value in enumerate(voc_values, 1):
            competitors.cell(31 + index, column, value)

    raw = workbook["MCP原始数据"]
    raw.cell(6, 1, "{SNAPSHOT_MONTH}")
    raw.cell(6, 2, "{NODE_ID_PATH}")
    clear_range(raw, 6, 6, 3, 12)
    for index, row in enumerate(range(10, 22), 1):
        raw.cell(row, 1, f"{{TREND_MONTH_{index:02d}}}")
        clear_range(raw, row, row, 2, 7)
        raw.cell(row, 8, f"{{TREND_EVIDENCE_ID_{index:02d}}}")
    for index, row in enumerate(range(25, 29), 1):
        raw.cell(row, 1, f"{{COMPETITOR_ASIN_{index}}}")
        raw.cell(row, 2, "user" if index <= 3 else "system")
        raw.cell(row, 3, f"{{BRAND_{index}}}")
        clear_range(raw, row, row, 4, 7)
        for column, field in enumerate(("FULFILLMENT", "LISTING_DATE", "NODE_ID_PATH", "ROLE"), 8):
            raw.cell(row, column, f"{{{field}_{index}}}")
        raw.cell(row, 12, None)
        raw.cell(row, 13, f"{{EVIDENCE_IDS_{index}}}")
    for index, row in enumerate(range(32, 39), 1):
        raw.cell(row, 1, f"{{KEYWORD_{index}}}")
        clear_range(raw, row, row, 2, 7)
        raw.cell(row, 8, "{MCP_KEYWORD_TOOL}")
        raw.cell(row, 9, f"{{KEYWORD_EVIDENCE_ID_{index}}}")

    workbook.properties.title = "Product-Planning V1 Data-Driven Reusable Template"
    workbook.properties.subject = "Sanitized seven-sheet Amazon product-planning template"
    workbook.properties.creator = ""
    workbook.properties.lastModifiedBy = ""
    workbook.properties.keywords = ""
    workbook.properties.description = "Runtime values are populated only from validated evidence.json."


def verify(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != renderer.SHEET_ORDER:
        raise ValueError(f"Unexpected sheet order: {workbook.sheetnames}")
    expected_charts = {"市场分析": 7, "竞品分析与优化策略": 1, "ABA排名【季度】": 0}
    for sheet_name, count in expected_charts.items():
        if len(workbook[sheet_name]._charts) != count:
            raise ValueError(f"{sheet_name} chart count is not {count}")
    if workbook["意向产品"]._images:
        raise ValueError("Reusable template must not contain a product image")
    aba_sheet = workbook["ABA排名【季度】"]
    if any(cell.value is not None for row in aba_sheet.iter_rows() for cell in row):
        raise ValueError("ABA manual-input sheet must be empty")
    if aba_sheet.merged_cells.ranges or aba_sheet._images:
        raise ValueError("ABA manual-input sheet contains residual objects")
    text = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    blocked = [
        r"\bB0[A-Z0-9]{8}\b",
        r"AIza[0-9A-Za-z_-]{20,}",
        r"github_pat_",
        r"secret-key=",
        r"[A-Za-z]:\\Users\\",
    ]
    for pattern in blocked:
        if re.search(pattern, text, re.IGNORECASE):
            raise ValueError(f"Reusable template contains blocked content: {pattern}")
    with zipfile.ZipFile(path) as archive:
        for member in archive.namelist():
            if member.endswith(".xml"):
                ElementTree.fromstring(archive.read(member))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("templates/product_planning_standard_template_V1_beautified.xlsx"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("templates/product_planning_standard_template_V1_data_driven.xlsx"),
    )
    args = parser.parse_args()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(args.source)
    if workbook.sheetnames != renderer.SHEET_ORDER:
        raise ValueError(f"Source template sheet order mismatch: {workbook.sheetnames}")
    data = build_placeholder_evidence()
    renderer.render_intent(workbook["意向产品"], data)
    renderer.render_market(workbook["市场分析"], data)
    old_competitor = workbook["竞品分析与优化策略"]
    competitor_index = workbook.index(old_competitor)
    workbook.remove(old_competitor)
    competitor_sheet = workbook.create_sheet("竞品分析与优化策略", competitor_index)
    renderer.render_competitors(competitor_sheet, data)
    renderer.render_swot(workbook["SWOT分析"], data)
    old_aba = workbook["ABA排名【季度】"]
    aba_index = workbook.index(old_aba)
    workbook.remove(old_aba)
    aba_sheet = workbook.create_sheet("ABA排名【季度】", aba_index)
    renderer.render_aba(aba_sheet, data)
    renderer.render_raw(workbook["MCP原始数据"], data)
    renderer.render_sources(workbook["数据来源"], data)
    scrub_runtime_values(workbook)
    workbook.save(args.output)
    verify(args.output)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
