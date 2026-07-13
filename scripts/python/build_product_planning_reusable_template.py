"""Build a clean reusable Product-Planning V1 meeting template with the approved seed-competitor layout."""

from __future__ import annotations

import argparse
import shutil
import zipfile
from copy import copy
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


NAVY = "1F4E78"
BLUE = "5B9BD5"
LIGHT_BLUE = "D9EAF7"
LIGHTER_BLUE = "EDF4FB"
YELLOW = "FFF2CC"
RED = "F4CCCC"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="B7C9D6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_title(ws, title: str) -> None:
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(name="Microsoft YaHei", size=15, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    ws.row_dimensions[1].height = 28


def header(ws, row: int, labels: list[str]) -> None:
    for column, label in enumerate(labels, 1):
        cell = ws.cell(row, column, label)
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def fill_row(ws, row: int, values: list[str], fill: str = WHITE, height: int = 32) -> None:
    for column, value in enumerate(values, 1):
        cell = ws.cell(row, column, value)
        cell.font = Font(name="Microsoft YaHei", size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = height


def merged_note(ws, row: int, text: str, fill: str = YELLOW, height: int = 30) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row, 1, text)
    cell.font = Font(name="Microsoft YaHei", size=10, italic=True)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER
    ws.row_dimensions[row].height = height


def merged_content(ws, row: int, label: str, placeholder: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row, 1, f"{label}：{placeholder}")
    cell.font = Font(name="Microsoft YaHei", size=10)
    cell.fill = PatternFill("solid", fgColor=LIGHTER_BLUE)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = BORDER
    ws.row_dimensions[row].height = 32


def write_competitor_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = NAVY
    for column, width in {"A": 41.375, "B": 33.875, "C": 29.625, "D": 30.0}.items():
        ws.column_dimensions[column].width = width
    set_title(ws, "{{PRODUCT_DIRECTION}} 种子竞品分析与优化策略")
    merged_note(
        ws,
        2,
        "输入规则：命令提供 3 个用户指定 ASIN；经 asin_detail 校验后，仅补充 1 个系统样本。所有实时字段来自 MCP，Amazon 前端图片/截图只作为辅助证据。",
        LIGHT_BLUE,
        34,
    )

    ws.merge_cells("A4:D4")
    ws["A4"] = "基础参数对比（当前 MCP 字段）"
    ws["A4"].font = Font(name="Microsoft YaHei", size=12, bold=True)
    ws["A4"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["A4"].alignment = Alignment(vertical="center")
    ws["A4"].border = BORDER
    ws.row_dimensions[4].height = 22
    header(ws, 5, ["维度", "种子竞品 1", "种子竞品 2", "种子竞品 3"])
    compare_rows = [
        ["ASIN", "{{ASIN_1}}", "{{ASIN_2}}", "{{ASIN_3}}"],
        ["品牌", "{{BRAND_1}}", "{{BRAND_2}}", "{{BRAND_3}}"],
        ["上架日", "{{LISTING_DATE_1}}", "{{LISTING_DATE_2}}", "{{LISTING_DATE_3}}"],
        ["类目路径", "{{NODE_LABEL_PATH_1}}", "{{NODE_LABEL_PATH_2}}", "{{NODE_LABEL_PATH_3}}"],
        ["价格", "{{PRICE_1}}", "{{PRICE_2}}", "{{PRICE_3}}"],
        ["评分", "{{RATING_1}}", "{{RATING_2}}", "{{RATING_3}}"],
        ["评分数 / 评论数", "{{RATINGS_AND_REVIEWS_1}}", "{{RATINGS_AND_REVIEWS_2}}", "{{RATINGS_AND_REVIEWS_3}}"],
        ["LQS", "{{LQS_1}}", "{{LQS_2}}", "{{LQS_3}}"],
        ["Amazon Choice", "{{AMAZON_CHOICE_1}}", "{{AMAZON_CHOICE_2}}", "{{AMAZON_CHOICE_3}}"],
        ["EBC + 视频", "{{EBC_AND_VIDEO_1}}", "{{EBC_AND_VIDEO_2}}", "{{EBC_AND_VIDEO_3}}"],
        ["变体", "{{VARIATIONS_1}}", "{{VARIATIONS_2}}", "{{VARIATIONS_3}}"],
        ["子类 / 主类 BSR", "{{BSR_1}}", "{{BSR_2}}", "{{BSR_3}}"],
        ["重量", "{{WEIGHT_1}}", "{{WEIGHT_2}}", "{{WEIGHT_3}}"],
        ["月销量快照", "{{MONTHLY_UNITS_1}}", "{{MONTHLY_UNITS_2}}", "{{MONTHLY_UNITS_3}}"],
    ]
    for row, values in enumerate(compare_rows, 6):
        fill_row(ws, row, values, LIGHTER_BLUE if row % 2 == 0 else WHITE, 30)

    ws.merge_cells("A21:D21")
    ws["A21"] = "核心差异化信号对比（Amazon 前端 + asin_detail.features + 评论主题）"
    ws["A21"].font = Font(name="Microsoft YaHei", size=12, bold=True)
    ws["A21"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["A21"].alignment = Alignment(vertical="center")
    ws["A21"].border = BORDER
    ws.row_dimensions[21].height = 22
    header(ws, 22, ["信号类型", "种子竞品 1", "种子竞品 2", "种子竞品 3"])
    signal_rows = [
        ["品牌信号", "{{BRAND_SIGNAL_1_OR_NA}}", "{{BRAND_SIGNAL_2_OR_NA}}", "{{BRAND_SIGNAL_3_OR_NA}}"],
        ["专业信号", "{{PROFESSIONAL_SIGNAL_1}}", "{{PROFESSIONAL_SIGNAL_2}}", "{{PROFESSIONAL_SIGNAL_3}}"],
        ["设备兼容", "{{DEVICE_COMPATIBILITY_1}}", "{{DEVICE_COMPATIBILITY_2}}", "{{DEVICE_COMPATIBILITY_3}}"],
        ["场景定位", "{{SCENE_POSITIONING_1}}", "{{SCENE_POSITIONING_2}}", "{{SCENE_POSITIONING_3}}"],
        ["材质 / 便携", "{{MATERIAL_PORTABILITY_1}}", "{{MATERIAL_PORTABILITY_2}}", "{{MATERIAL_PORTABILITY_3}}"],
        ["融合方向启示", "{{DIRECTION_INSIGHT_1}}", "{{DIRECTION_INSIGHT_2}}", "{{DIRECTION_INSIGHT_3}}"],
    ]
    for row, values in enumerate(signal_rows, 23):
        fill_row(ws, row, values, LIGHTER_BLUE if row % 2 else WHITE, 38)
    merged_note(ws, 29, "图片读取规则：产品图和前端截图中的型号、售价、月销、评分等需标记为“截图快照”；实时分析字段优先使用本次 MCP 返回。没有接口字段时填写 N/A。", YELLOW, 36)

    start = 31
    for index in range(1, 5):
        ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=4)
        title = ws.cell(start, 1, f"ASIN：{{{{COMPETITOR_ASIN_{index}}}}}｜{'用户指定种子' if index <= 3 else '系统补充样本'}")
        title.font = Font(name="Microsoft YaHei", size=12, bold=True, color=WHITE)
        title.fill = PatternFill("solid", fgColor=NAVY)
        title.alignment = Alignment(vertical="center")
        title.border = BORDER
        ws.row_dimensions[start].height = 24
        merged_content(ws, start + 1, "基础快照", f"{{{{BRAND_{index}}}}}；售价 {{{{PRICE_{index}}}}}；月销量 {{{{MONTHLY_UNITS_{index}_OR_NA}}}}；评分 {{{{RATING_{index}}}}} / {{{{RATINGS_{index}}}}}；{{{{FULFILLMENT_{index}}}}}")
        merged_content(ws, start + 2, "用户画像", f"{{{{PERSONA_FROM_LISTING_AND_REVIEWS_{index}}}}}")
        merged_content(ws, start + 3, "核心场景", f"{{{{SCENARIO_FROM_LISTING_AND_REVIEWS_{index}}}}}")
        merged_content(ws, start + 4, "前端图片 / 截图信息", f"{{{{AMAZON_FRONTEND_IMAGE_OR_SCREENSHOT_INSIGHT_{index}}}}}")
        header(ws, start + 6, ["核心卖点（Listing）", "好评关键词（4-5 星）", "差评痛点（1-3 星）", "对标方向"])
        fill_row(
            ws,
            start + 7,
            [
                f"{{{{LISTING_FEATURES_{index}}}}}",
                f"{{{{POSITIVE_REVIEW_THEMES_{index}}}}}",
                f"{{{{NEGATIVE_REVIEW_THEMES_{index}}}}}",
                f"{{{{PRODUCT_MANAGER_ACTION_{index}}}}}",
            ],
            WHITE if index % 2 else LIGHTER_BLUE,
            110,
        )
        merged_note(ws, start + 8, f"证据ID：{{{{EVIDENCE_IDS_{index}}}}}；空字段：{{{{DATA_GAPS_{index}_OR_NA}}}}", YELLOW, 28)
        start += 10

    merged_note(ws, start, "结论规则：最多输出 3 条产品经理动作，分别覆盖样品验证、Listing 表达、定价或投放；结论必须回查到 ASIN 与证据ID。", RED, 34)
    ws.freeze_panes = "A5"


def clean_template_metadata(workbook) -> None:
    workbook.properties.title = "Product Planning Reusable Meeting Template"
    workbook.properties.subject = "Sanitized reusable product planning workbook"
    workbook.properties.creator = ""
    workbook.properties.lastModifiedBy = ""
    workbook.properties.keywords = ""
    workbook.properties.description = "Reusable template with placeholders only."


def verify(path: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    expected = ["意向产品", "市场分析", "竞品分析与优化策略", "SWOT分析", "ABA排名【季度】", "MCP原始数据", "数据来源"]
    if workbook.sheetnames != expected:
        raise ValueError(f"Unexpected worksheet order: {workbook.sheetnames}")
    sheet = workbook["竞品分析与优化策略"]
    if sheet["A31"].value != "ASIN：{{COMPETITOR_ASIN_1}}｜用户指定种子":
        raise ValueError("Seed competitor card placeholder is missing")
    if sheet["A61"].value != "ASIN：{{COMPETITOR_ASIN_4}}｜系统补充样本":
        raise ValueError("System supplement card placeholder is missing")
    with zipfile.ZipFile(path) as archive:
        for member in archive.namelist():
            if member.endswith(".xml"):
                ElementTree.fromstring(archive.read(member))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=Path("templates/product_planning_meeting_template.xlsx"))
    parser.add_argument("--output", type=Path, default=Path("templates/product_planning_reusable_template_V1.xlsx"))
    args = parser.parse_args()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(args.source, args.output)
    workbook = load_workbook(args.output)
    original = next(sheet for sheet in workbook.worksheets if "品分析与优化策略" in sheet.title)
    index = workbook.index(original)
    workbook.remove(original)
    competitor = workbook.create_sheet("竞品分析与优化策略", index)
    write_competitor_sheet(competitor)
    clean_template_metadata(workbook)
    workbook.save(args.output)
    verify(args.output)
    print(args.output)


if __name__ == "__main__":
    main()
