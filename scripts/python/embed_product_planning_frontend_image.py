"""Embed the first seed ASIN's Amazon frontend image into Product Planning V1."""

from __future__ import annotations

import argparse
import copy
import shutil
import zipfile
from datetime import date
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels


def column_width_to_pixels(width: float | None) -> int:
    """Approximate Excel column width in pixels using the default font metrics."""
    effective = 8.43 if width is None else width
    return max(1, int(effective * 7 + 5))


def image_anchor(sheet, cell_ref: str, width: int, height: int) -> OneCellAnchor:
    column_name, row = coordinate_from_string(cell_ref)
    column = column_index_from_string(column_name)
    cell_width = column_width_to_pixels(sheet.column_dimensions[column_name].width)
    row_height = points_to_pixels(sheet.row_dimensions[row].height or 15)
    x_offset = max(0, (cell_width - width) // 2)
    y_offset = max(0, (row_height - height) // 2)
    marker = AnchorMarker(
        col=column - 1,
        colOff=pixels_to_EMU(x_offset),
        row=row - 1,
        rowOff=pixels_to_EMU(y_offset),
    )
    extent = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
    return OneCellAnchor(_from=marker, ext=extent)


def remove_image_at(sheet, cell_ref: str) -> None:
    column_name, row = coordinate_from_string(cell_ref)
    target_col = column_index_from_string(column_name) - 1
    target_row = row - 1
    retained = []
    for image in sheet._images:
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is not None and marker.col == target_col and marker.row == target_row:
            continue
        retained.append(image)
    sheet._images = retained


def copy_row_style(sheet, source_row: int, target_row: int, max_column: int) -> None:
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.alignment = copy.copy(source.alignment)
        target.border = copy.copy(source.border)
        target.fill = copy.copy(source.fill)
        target.font = copy.copy(source.font)


def register_source(workbook, asin: str, marketplace: str, source_url: str) -> None:
    sheet = workbook["数据来源"]
    target_row = None
    for row in range(4, sheet.max_row + 1):
        if sheet.cell(row, 1).value == "Amazon 前端首图":
            target_row = row
            break
    if target_row is None:
        target_row = sheet.max_row + 1
        copy_row_style(sheet, max(4, target_row - 1), target_row, max(5, sheet.max_column))
    values = [
        "Amazon 前端首图",
        f"第一个种子 ASIN {asin}；{marketplace}；抓取日期 {date.today().isoformat()}",
        "Amazon 商品详情页首图",
        "前端图片为抓取时点快照，Listing 更新后可能变化",
        "意向产品",
    ]
    for column, value in enumerate(values, 1):
        sheet.cell(target_row, column, value)
    if source_url:
        sheet.cell(target_row, 2).comment = Comment(f"公开来源：{source_url}", "Codex")


def embed_image(
    workbook_path: Path,
    image_path: Path,
    output_path: Path,
    asin: str,
    marketplace: str,
    source_url: str,
    sheet_name: str = "意向产品",
    cell_ref: str = "A5",
) -> None:
    if workbook_path.resolve() != output_path.resolve():
        shutil.copy2(workbook_path, output_path)

    workbook = load_workbook(output_path)
    sheet = workbook[sheet_name]
    column_name, row = coordinate_from_string(cell_ref)
    cell_width = column_width_to_pixels(sheet.column_dimensions[column_name].width)
    cell_height = points_to_pixels(sheet.row_dimensions[row].height or 15)

    picture = ExcelImage(str(image_path))
    scale = min((cell_width - 8) / picture.width, (cell_height - 8) / picture.height)
    picture.width = max(1, int(picture.width * scale))
    picture.height = max(1, int(picture.height * scale))
    picture.anchor = image_anchor(sheet, cell_ref, picture.width, picture.height)

    remove_image_at(sheet, cell_ref)
    sheet[cell_ref] = None
    sheet.add_image(picture)
    register_source(workbook, asin, marketplace, source_url)

    temporary = output_path.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    temporary.replace(output_path)

    with zipfile.ZipFile(output_path) as archive:
        for member in archive.namelist():
            if member.endswith(".xml"):
                ElementTree.fromstring(archive.read(member))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("image", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--asin", required=True)
    parser.add_argument("--marketplace", default="US")
    parser.add_argument("--source-url", default="")
    parser.add_argument("--sheet", default="意向产品")
    parser.add_argument("--cell", default="A5")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = args.output or args.workbook
    embed_image(
        args.workbook,
        args.image,
        output,
        args.asin,
        args.marketplace,
        args.source_url,
        args.sheet,
        args.cell,
    )
    print(output)


if __name__ == "__main__":
    main()

