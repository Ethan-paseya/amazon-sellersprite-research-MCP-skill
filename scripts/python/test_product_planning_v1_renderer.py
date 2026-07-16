#!/usr/bin/env python3
"""Boundary regression tests for the Product-Planning V1 renderer."""

from __future__ import annotations

import sys
import unittest
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent))
import build_product_planning_data_driven_template as template_builder
import render_product_planning_v1 as renderer


def evidence_with_missing_supplement() -> dict:
    data = template_builder.build_placeholder_evidence()
    data["command"]["systemSupplement"] = "N/A"
    data["competitors"][3] = {
        "asin": "N/A",
        "source": "system",
        "brand": "N/A",
        "title": "N/A",
        "price": "N/A",
        "rating": "N/A",
        "ratings": "N/A",
        "reviews": "N/A",
        "fulfillment": "N/A",
        "listingDate": "N/A",
        "categoryPath": "N/A",
        "nodeIdPath": "N/A",
        "role": "系统候选池无合格样本",
        "productAction": "扩大候选池后重试，不补造竞品数据",
        "features": "N/A",
        "positiveThemes": "N/A",
        "negativeThemes": "N/A",
        "priority": "P1",
        "salesSnapshot": "N/A",
        "evidenceIds": [],
        "gapReason": "MCP 候选池无通过相关性校验的第四竞品",
        "imageUrl": "N/A",
    }
    return data


class ProductPlanningV1RendererTests(unittest.TestCase):
    def test_missing_system_competitor_renders_as_na(self) -> None:
        data = evidence_with_missing_supplement()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "竞品分析与优化策略"

        renderer.render_competitors(sheet, data)

        self.assertEqual("N/A", sheet["A9"].value)
        self.assertEqual("N/A", sheet["E9"].value)
        self.assertIsNone(sheet["K16"].value)
        self.assertIsNone(sheet["L16"].value)
        self.assertIn("N/A", renderer.render_markdown(data, None))

    def test_public_single_brace_placeholder_is_detected(self) -> None:
        self.assertIsNotNone(renderer.PLACEHOLDER_RE.search("{PRODUCT_DIRECTION}"))

    def test_aba_sheet_is_reset_to_blank_manual_input(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "old ABA value"
        renderer.render_aba(sheet, {})
        self.assertTrue(all(cell.value is None for row in sheet.iter_rows() for cell in row))
        self.assertTrue(sheet.sheet_view.showGridLines)


if __name__ == "__main__":
    unittest.main()
