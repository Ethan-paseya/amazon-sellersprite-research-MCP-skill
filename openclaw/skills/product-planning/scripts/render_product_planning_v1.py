"""Render Product-Planning V1 Markdown and Excel from normalized evidence JSON."""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import urllib.request
import zipfile
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_ORDER = [
    "意向产品",
    "市场分析",
    "竞品分析与优化策略",
    "SWOT分析",
    "ABA排名【季度】",
    "MCP原始数据",
    "数据来源",
]
NAVY = "1F4E78"
BLUE = "5B9BD5"
LIGHT_BLUE = "DDEBF7"
PALE_BLUE = "EAF2F8"
YELLOW = "FFF2CC"
BRIGHT_YELLOW = "FFFF00"
LIGHT_GRAY = "F2F2F2"
MID_GRAY = "BFBFBF"
LIGHT_RED = "FCE4D6"
LIGHT_GREEN = "E2F0D9"
WHITE = "FFFFFF"
TEXT = "203040"
RED = "FF0000"
THIN = Side(style="thin", color="AAB8C2")
SECRET_RE = re.compile(
    r"(?:github_pat_[A-Za-z0-9_]+|AIza[0-9A-Za-z_-]{20,}|secret-key=[^&\s\"]+|"
    r"(?:api[_-]?key|token|secret)\s*[:=]\s*[\"']?[A-Za-z0-9._-]{16,})",
    re.IGNORECASE,
)
ABS_PATH_RE = re.compile(r"(?:[A-Za-z]:\\Users\\|/Users/|/home/)", re.IGNORECASE)
PLACEHOLDER_RE = re.compile(
    r"\{\{[^{}]+\}\}|\{[A-Z][A-Z0-9_]+\}|\bTODO\b|\bPLACEHOLDER\b",
    re.IGNORECASE,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("evidence", type=Path)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--qa", type=Path)
    parser.add_argument("--validation", type=Path)
    return parser.parse_args()


def load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def number_or_none(value: object) -> float | None:
    """Return a finite number or None for MCP gaps such as N/A and empty values."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str) and value.strip().upper() in {"", "N/A", "NA", "NULL", "NONE", "-"}:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def number_or_zero(value: object) -> float:
    number = number_or_none(value)
    return number if number is not None else 0.0


def text_or_na(value: object) -> str:
    if value is None:
        return "N/A"
    text = str(value).strip()
    return text or "N/A"


def fixed_text_items(values: object, size: int) -> list[str]:
    items = [text_or_na(value) for value in values] if isinstance(values, list) else []
    return (items + ["N/A"] * size)[:size]


def format_money(value: object) -> str:
    number = number_or_none(value)
    return f"${number:,.2f}" if number is not None else "N/A"


def format_number(value: object, decimals: int = 0) -> str:
    number = number_or_none(value)
    if number is None:
        return "N/A"
    return f"{number:,.{decimals}f}"


def format_percent(value: object, decimals: int = 2) -> str:
    number = number_or_none(value)
    return f"{number:.{decimals}%}" if number is not None else "N/A"


def reset_sheet(sheet) -> None:
    for merged in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged))
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None
            cell.comment = None
    sheet._charts = []
    sheet._images = []
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = NAVY
    sheet.freeze_panes = None


def style_range(
    sheet,
    cell_range: str,
    fill: str = WHITE,
    font_color: str = TEXT,
    bold: bool = False,
    size: float = 10.5,
    horizontal: str = "left",
) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name="微软雅黑", size=size, bold=bold, color=font_color)
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def merge_title(sheet, cell_range: str, value: str, fill: str = NAVY, size: float = 15) -> None:
    sheet.merge_cells(cell_range)
    cell = sheet[cell_range.split(":")[0]]
    cell.value = value
    style_range(sheet, cell_range, fill, WHITE if fill == NAVY else TEXT, True, size, "left")


def merge_note(sheet, cell_range: str, value: str, fill: str = YELLOW, size: float = 9.5) -> None:
    sheet.merge_cells(cell_range)
    cell = sheet[cell_range.split(":")[0]]
    cell.value = value
    style_range(sheet, cell_range, fill, TEXT, False, size, "left")


def header(sheet, row: int, start: int, end: int, values: list[str] | None = None) -> None:
    if values:
        for column, value in enumerate(values, start):
            sheet.cell(row, column, value)
    style_range(
        sheet,
        f"{get_column_letter(start)}{row}:{get_column_letter(end)}{row}",
        NAVY,
        WHITE,
        True,
        10.5,
        "center",
    )
    sheet.row_dimensions[row].height = 28


def alternating(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        style_range(
            sheet,
            f"{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}",
            LIGHT_GRAY if row % 2 == 0 else WHITE,
        )


def setup_page(sheet, area: str) -> None:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    sheet.print_area = area


def add_dual_chart(
    sheet,
    anchor: str,
    min_row: int,
    max_row: int,
    category_col: int,
    bar_col: int,
    line_col: int,
    title: str,
    bar_title: str,
    line_title: str,
    width: float = 14.0,
    height: float = 7.2,
) -> None:
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = title
    bar.y_axis.title = bar_title
    bar.add_data(Reference(sheet, min_col=bar_col, min_row=min_row - 1, max_row=max_row), titles_from_data=True)
    bar.set_categories(Reference(sheet, min_col=category_col, min_row=min_row, max_row=max_row))
    line = LineChart()
    line.add_data(Reference(sheet, min_col=line_col, min_row=min_row - 1, max_row=max_row), titles_from_data=True)
    line.y_axis.title = line_title
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"
    bar += line
    bar.width = width
    bar.height = height
    sheet.add_chart(bar, anchor)


def add_distribution_chart(sheet, anchor: str, min_row: int, max_row: int, title: str, start_col: int) -> None:
    add_dual_chart(
        sheet,
        anchor,
        min_row,
        max_row,
        start_col,
        start_col + 1,
        start_col + 2,
        title,
        "产品数",
        "销量占比",
        13.2,
        6.8,
    )


def download_image(url: str) -> bytes | None:
    if not url.startswith("https://"):
        return None
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            return response.read()
    except Exception:
        return None


def render_intent(sheet, data: dict, preserved_image: bytes | None = None) -> bool:
    reset_sheet(sheet)
    product = data["productDirection"]
    first = data["competitors"][0]
    snapshot = data["market"]["snapshot"]
    verified_functions = fixed_text_items(product.get("verifiedFunctions"), 4)
    unmet_needs = fixed_text_items(product.get("unmetNeeds"), 3)
    merge_title(sheet, "A1:J1", f"产品意向｜{product['name']}")
    merge_note(
        sheet,
        "A2:J2",
        "数据提炼来源：Amazon 前端产品介绍/主图 + SellerSprite MCP 竞品与评论；功能一至四为已验证卖点，功能五至七为低星评论中可转化的未满足需求。",
        LIGHT_BLUE,
    )
    merge_note(sheet, "A3:J3", f"意向1｜{product['name']}", BLUE, 12)
    headers = [
        "参考图片",
        "价格定位",
        "客户人群",
        "功能一",
        "功能二",
        "功能三",
        "功能四",
        "功能五(未满足需求)",
        "功能六(未满足需求)",
        "功能七(未满足需求)",
    ]
    for col, value in enumerate(headers, 1):
        sheet.cell(4, col, value)
    style_range(sheet, "A4:C4", MID_GRAY, TEXT, True, 10.5, "center")
    style_range(sheet, "D4:G4", MID_GRAY, RED, True, 10.5, "center")
    style_range(sheet, "H4:J4", BRIGHT_YELLOW, RED, True, 10.5, "center")
    values = [
        "",
        f"{text_or_na(product.get('targetPrice'))}\n主节点均价 {format_money(snapshot.get('avgPrice'))}",
        text_or_na(product.get("targetAudience")),
        *verified_functions,
        *unmet_needs,
    ]
    for col, value in enumerate(values, 1):
        sheet.cell(5, col, value)
    style_range(sheet, "A5:C5", WHITE)
    style_range(sheet, "D5:G5", PALE_BLUE)
    style_range(sheet, "H5:J5", LIGHT_RED)
    image_data = preserved_image or download_image(first.get("imageUrl", ""))
    embedded = False
    if image_data:
        image = Image(BytesIO(image_data))
        image.width = 145
        image.height = 108
        image.anchor = "A5"
        sheet.add_image(image)
        embedded = True
    else:
        sheet["A5"] = "N/A（首个 ASIN 主图未能嵌入）"
    merge_note(
        sheet,
        "A7:J7",
        "填写口径：参考图使用第一款种子 ASIN 已核验主图；价格为小样测试带而非利润结论；功能一至四来自 Listing/高星评论，功能五至七来自低星 VOC；工程、供应链、成本与专利仍需单独验证。",
    )
    widths = [18, 17, 22, 20, 22, 20, 20, 22, 22, 22]
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    for row, height in {1: 28, 2: 36, 3: 26, 4: 42, 5: 115, 6: 8, 7: 40}.items():
        sheet.row_dimensions[row].height = height
    sheet.freeze_panes = "D5"
    setup_page(sheet, "A1:J7")
    return embedded


def write_distribution_table(sheet, start_row: int, start_col: int, items: list[dict]) -> int:
    header(sheet, start_row, start_col, start_col + 2, ["区间", "产品数", "销量占比"])
    for row, item in enumerate(items, start_row + 1):
        sheet.cell(row, start_col, item["label"])
        sheet.cell(row, start_col + 1, item.get("products", "N/A"))
        sheet.cell(row, start_col + 2, item["share"])
        sheet.cell(row, start_col + 2).number_format = "0.00%"
    end = start_row + len(items)
    alternating(sheet, start_row + 1, end, start_col, start_col + 2)
    return end


def render_market(sheet, data: dict) -> None:
    reset_sheet(sheet)
    selected = data["categorySelection"]["selected"]
    market = data["market"]
    product = data["productDirection"]
    template_mode = bool(data.get("render", {}).get("templateMode"))
    keywords = data["keywords"]["items"][:7]
    trend = market["monthlyTrend"][:12]
    snapshot = market["snapshot"]
    dist = market["distributions"]
    concentration = market["concentration"]

    def dominant(items: list[dict]) -> dict:
        return max(items, key=lambda item: number_or_zero(item.get("share")), default={"label": "N/A", "share": 0})

    units = [number_or_zero(item.get("avgUnits")) for item in trend]
    prices = [number for item in trend if (number := number_or_none(item.get("avgPrice"))) is not None] or [0.0]
    avg_units = sum(units) / len(units) if units else 0
    peak = max(trend, key=lambda item: number_or_zero(item.get("avgUnits")), default={"month": "N/A", "avgUnits": 0})
    top_keyword = max(keywords, key=lambda item: number_or_zero(item.get("searches")), default={"keyword": "N/A", "searches": 0})
    top_purchase = max(keywords, key=lambda item: number_or_zero(item.get("purchaseRate")), default={"keyword": "N/A", "purchaseRate": 0})
    best_supply = max(keywords, key=lambda item: number_or_zero(item.get("supplyDemandRatio")), default={"keyword": "N/A", "supplyDemandRatio": 0})
    age_items = dist["listingAge"]["items"][:9]
    year_items = dist["listingYear"]["items"][:10]
    rating_count_items = dist["ratingCount"]["items"][:7]
    rating_items = dist["rating"]["items"][:7]
    price_items = dist["price"]["items"][:9]
    top_age = dominant(age_items)
    top_year = dominant(year_items)
    top_rating_count = dominant(rating_count_items)
    top_rating = dominant(rating_items)
    top_price = dominant(price_items)

    if template_mode:
        trend_comment = "产品经理点评：运行时根据连续 12 个月节点销量、销售额、价格和新品效率自动生成。"
        keyword_comment = "产品经理点评：运行时根据搜索量、购买率、PPC、商品数和供需比自动生成。"
        age_comment = "产品经理点评：运行时根据节点上架时长分布自动生成。"
        year_comment = "产品经理点评：运行时根据节点上架年份分布自动生成。"
        ratings_comment = "产品经理点评：运行时根据评分数区间及销量占比自动生成。"
        rating_comment = "产品经理点评：运行时根据评分值区间及销量占比自动生成。"
        total_comment = "市场分析总评：运行时根据 evidence.json 的节点快照、趋势、分布和集中度自动生成。"
    else:
        trend_comment = (
            f"产品经理点评：\n1、连续 {len(trend)} 个月平均月销约 {avg_units:,.0f} 件，峰值为 {peak['month']} 的 {number_or_zero(peak.get('avgUnits')):,.0f} 件。\n"
            f"2、月均价范围为 ${min(prices):.2f}-${max(prices):.2f}；最新月新品均销 {snapshot.get('newAvgUnits', 'N/A')}，类目均销 {snapshot.get('avgUnits', 'N/A')}。\n"
            f"3、需求接口处理：{market.get('demandTrend', {}).get('handling', '按接口真实返回记录')}。"
        )
        keyword_comment = (
            f"产品经理点评：\n1、搜索量最高词为 {top_keyword['keyword']}（{format_number(top_keyword.get('searches'))}）。\n"
            f"2、购买率最高词为 {top_purchase['keyword']}（{format_percent(top_purchase.get('purchaseRate'))}）。\n"
            f"3、供需比最高词为 {best_supply['keyword']}（{format_number(best_supply.get('supplyDemandRatio'), 2)}），仍需结合相关性与 PPC 判断切入优先级。"
        )
        age_comment = f"产品经理点评：销量占比最高的上架时长区间为 {top_age['label']}（{format_percent(top_age.get('share'))}）；据此判断成熟产品与新品的相对效率。"
        year_comment = f"产品经理点评：销量占比最高的上架年份为 {top_year['label']}（{format_percent(top_year.get('share'))}）；新品进入判断不得只看产品数量。"
        ratings_comment = f"产品经理点评：评分数区间 {top_rating_count['label']} 承接最高销量占比 {format_percent(top_rating_count.get('share'))}，作为评价壁垒基准。"
        rating_comment = f"产品经理点评：评分值区间 {top_rating['label']} 承接最高销量占比 {format_percent(top_rating.get('share'))}，样品品质目标应至少对齐主流区间。"
        total_comment = (
            f"市场分析总评（产品经理）：\n1、{snapshot['month']} 节点样本 {format_number(snapshot.get('products'))} 个、均价 {format_money(snapshot.get('avgPrice'))}、平均月销 {format_number(snapshot.get('avgUnits'))}、平均评分 {format_number(snapshot.get('avgRating'), 1)}、平均评分数 {format_number(snapshot.get('avgRatings'))}。\n"
            f"2、销量占比最高价格带为 {top_price['label']}（{format_percent(top_price.get('share'))}）；本项目测试价格 {product['targetPrice']} 必须由可验证差异支撑。\n"
            f"3、品牌 CR5 月销量 {format_percent(concentration.get('brandCr5Units'))}，需结合评价壁垒与产品差异判断可进入性。"
        )
    merge_title(sheet, "A1:X1", f"类目：{selected['labelPath']}｜目标节点：{selected['nodeIdPath']}", BLUE, 13)
    merge_note(sheet, "A2:X2", "市场概况与规模：销售趋势、关键词、上架时间/年份、评分数、评分值和价格区间均使用目标叶子节点的 MCP 数据。", LIGHT_GRAY)
    merge_note(sheet, "A3:X3", f"类目选择依据：{selected['reason']}；例外说明：{selected['overrideReason']}", LIGHT_BLUE)

    merge_note(sheet, "A4:L4", "1. 销售趋势｜MCP：market_research_statistics", BLUE, 11)
    header(sheet, 5, 1, 4, ["月份", "月均销量", "月均销售额($)", "平均BSR"])
    for row, item in enumerate(trend, 6):
        sheet.cell(row, 1, item["month"])
        sheet.cell(row, 2, item["avgUnits"])
        sheet.cell(row, 3, item["avgRevenue"])
        sheet.cell(row, 4, item["avgBsr"])
    alternating(sheet, 6, 17, 1, 4)
    add_dual_chart(sheet, "E5", 6, 17, 1, 2, 3, "目标节点 12 个月销售趋势", "月均销量", "月均销售额($)")

    merge_note(sheet, "M4:X4", "2. 关键词｜MCP：keyword_miner / traffic_keyword", BLUE, 11)
    header(sheet, 5, 13, 17, ["关键词", "搜索量", "购买量", "PPC($)", "商品数"])
    for row, item in enumerate(keywords, 6):
        sheet.cell(row, 13, item["keyword"])
        sheet.cell(row, 14, item["searches"])
        sheet.cell(row, 15, item["purchases"])
        sheet.cell(row, 16, item["ppc"])
        sheet.cell(row, 17, item["products"])
    alternating(sheet, 6, 12, 13, 17)
    add_dual_chart(sheet, "R5", 6, 12, 13, 14, 15, "目标产品关键词", "月搜索量", "月购买量", 13.0, 7.2)
    merge_note(sheet, "A18:L20", trend_comment)
    merge_note(sheet, "M18:X20", keyword_comment)

    merge_note(sheet, "A22:L22", "3. 上架时间分布｜MCP：market_listing_date_distribution", BLUE, 11)
    merge_note(sheet, "M22:X22", "4. 上架年份分布｜MCP：market_listing_trend", BLUE, 11)
    end_age = write_distribution_table(sheet, 23, 1, age_items)
    end_year = write_distribution_table(sheet, 23, 13, year_items)
    add_distribution_chart(sheet, "E23", 24, end_age, "上架时间分布", 1)
    add_distribution_chart(sheet, "Q23", 24, end_year, "上架年份分布", 13)
    merge_note(sheet, "A34:L36", age_comment)
    merge_note(sheet, "M34:X36", year_comment)

    merge_note(sheet, "A38:L38", "5. 评分数分析｜MCP：market_ratings_distribution", BLUE, 11)
    merge_note(sheet, "M38:X38", "6. 评分值分布｜MCP：market_rating_distribution", BLUE, 11)
    end_ratings = write_distribution_table(sheet, 39, 1, rating_count_items)
    end_rating = write_distribution_table(sheet, 39, 13, rating_items)
    add_distribution_chart(sheet, "E39", 40, end_ratings, "评分数与销量占比", 1)
    add_distribution_chart(sheet, "Q39", 40, end_rating, "评分值与销量占比", 13)
    merge_note(sheet, "A47:L49", ratings_comment)
    merge_note(sheet, "M47:X49", rating_comment)

    merge_note(sheet, "A51:L51", "7. 价格区间｜MCP：market_price_distribution", BLUE, 11)
    end_price = write_distribution_table(sheet, 52, 1, price_items)
    add_distribution_chart(sheet, "E52", 53, end_price, "价格区间与销量占比", 1)
    merge_note(sheet, "A62:X65", total_comment, YELLOW, 10)
    for col in range(1, 25):
        sheet.column_dimensions[get_column_letter(col)].width = 12.5
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["M"].width = 27
    for row in (18, 19, 20, 34, 35, 36, 47, 48, 49, 62, 63, 64, 65):
        sheet.row_dimensions[row].height = 22
    sheet.freeze_panes = "A4"
    setup_page(sheet, "A1:X65")


def render_competitors(sheet, data: dict) -> None:
    reset_sheet(sheet)
    competitors = data.get("competitors", [])
    if len(competitors) != 4:
        raise ValueError("竞品分析需要固定 4 行：3 款用户种子 + 1 款系统补充或 N/A 缺口行")
    product = data["productDirection"]
    merge_title(sheet, "A1:L1", f"{product['name']}｜竞品分析与优化策略")
    merge_note(
        sheet,
        "A2:L2",
        "样本口径：3 款用户指定种子经 asin_detail 校验后保留；系统从候选池补充 1 款必要样本，用于覆盖头部评价、低价冲击或高配差异化。数值来自 MCP，卖点/VOC 只使用 Listing 与分星级评论。",
        LIGHT_BLUE,
    )
    merge_note(sheet, "A4:L4", "代表竞品表｜SellerSprite MCP + Amazon 前端卖点", PALE_BLUE, 11)
    cols = ["ASIN", "来源", "品牌", "标题", "售价", "评分", "评分数", "履约", "上架时间", "类目路径", "竞品角色", "产品经理动作"]
    header(sheet, 5, 1, 12, cols)
    for row, item in enumerate(competitors, 6):
        source = item.get("source")
        source_label = "用户种子" if source == "user" else ("系统补充" if source == "system" else "N/A")
        price = number_or_none(item.get("price"))
        rating = number_or_none(item.get("rating"))
        ratings = number_or_none(item.get("ratings"))
        values = [
            text_or_na(item.get("asin")), source_label, text_or_na(item.get("brand")), text_or_na(item.get("title")),
            price if price is not None else "N/A", rating if rating is not None else "N/A", ratings if ratings is not None else "N/A",
            text_or_na(item.get("fulfillment")), text_or_na(item.get("listingDate")), text_or_na(item.get("categoryPath")),
            text_or_na(item.get("role")), text_or_na(item.get("productAction")),
        ]
        for col, value in enumerate(values, 1):
            sheet.cell(row, col, value)
        sheet.cell(row, 5).number_format = '$0.00'
    alternating(sheet, 6, 9, 1, 12)
    for row in range(6, 10):
        sheet.row_dimensions[row].height = 78

    header(sheet, 12, 10, 12, ["ASIN", "售价（USD）", "评分数"])
    for row, item in enumerate(competitors, 13):
        sheet.cell(row, 10, text_or_na(item.get("asin")))
        sheet.cell(row, 11, number_or_none(item.get("price")))
        sheet.cell(row, 12, number_or_none(item.get("ratings")))
    add_dual_chart(sheet, "A12", 13, 16, 10, 11, 12, "代表竞品：售价与评分数", "售价（USD）", "评分数", 15.5, 8.0)

    merge_note(sheet, "A30:L30", "Listing / VOC 优化矩阵", PALE_BLUE, 11)
    header(sheet, 31, 1, 7, ["ASIN", "已验证卖点", "高星认可", "低星痛点", "可执行优化策略", "优先级", "证据ID"])
    for row, item in enumerate(competitors, 32):
        priority = item.get("priority") or ("P1" if "未返回" in str(item.get("negativeThemes", "")) else "P0")
        evidence_ids = item.get("evidenceIds") or []
        values = [
            text_or_na(item.get("asin")), text_or_na(item.get("features")), text_or_na(item.get("positiveThemes")),
            text_or_na(item.get("negativeThemes")), text_or_na(item.get("productAction")), text_or_na(priority),
            "; ".join(str(value) for value in evidence_ids) or "N/A",
        ]
        for col, value in enumerate(values, 1):
            sheet.cell(row, col, value)
        sheet.row_dimensions[row].height = 82
    alternating(sheet, 32, 35, 1, 7)
    actions = data.get("decision", {}).get("actions", [])[:3]
    merge_note(sheet, "A38:L41", "产品经理结论：\n" + "\n".join(f"{i + 1}、{item['priority']}：{item['action']}" for i, item in enumerate(actions)), YELLOW, 10)

    widths = [15, 12, 14, 45, 12, 10, 13, 12, 14, 42, 24, 45]
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A6"
    setup_page(sheet, "A1:L41")


def render_swot(sheet, data: dict) -> None:
    reset_sheet(sheet)
    merge_title(sheet, "A1:B1", "SWOT 分析｜仅重组前 3 个 Sheet 的已验证证据")
    header(sheet, 3, 1, 2, ["象限", "证据化结论"])
    market = data["market"]
    template_mode = bool(data.get("render", {}).get("templateMode"))
    if template_mode:
        rows = [(label, "{运行时从前三个 Sheet 的 evidenceId 自动生成}") for label in ("S 优势", "S 优势", "W 劣势", "W 劣势", "O 机会", "O 机会", "T 威胁", "T 威胁")]
    else:
        verified = data["productDirection"].get("verifiedFunctions", [])
        gaps = data.get("dataGaps", [])
        trend = market.get("monthlyTrend", [])
        avg_units = sum(number_or_zero(item.get("avgUnits")) for item in trend) / len(trend) if trend else 0
        keywords = data.get("keywords", {}).get("items", [])
        top_keyword = max(keywords, key=lambda item: number_or_zero(item.get("searches")), default={"keyword": "N/A", "searches": 0})
        rating_barrier = max(market["distributions"]["ratingCount"]["items"], key=lambda item: number_or_zero(item.get("share")), default={"label": "N/A", "share": 0})
        supplement = next((item for item in data["competitors"] if item.get("source") == "system"), data["competitors"][-1])
        ip_terms = "、".join(str(item.get("term")) for item in data.get("ipSignals", []) if item.get("term")) or "商标/专利"
        rows = [
            ("S 优势", "；".join(verified[:2]) or "种子竞品已验证的核心功能可被复用。"),
            ("S 优势", "；".join(item.get("positiveThemes", "") for item in data["competitors"][:2])),
            ("W 劣势", gaps[0]["description"] if gaps else "关键字段存在数据缺口。"),
            ("W 劣势", gaps[-1]["description"] if gaps else "工程、供应链和成本仍需独立验证。"),
            ("O 机会", f"类目连续 {len(trend)} 个月平均月销约 {avg_units:,.0f}；最高搜索词 {top_keyword['keyword']} 为 {format_number(top_keyword.get('searches'))}。"),
            ("O 机会", f"品牌 CR5 月销量 {format_percent(market['concentration'].get('brandCr5Units'))}；结合差异化功能判断进入空间。"),
            ("T 威胁", f"评分数区间 {rating_barrier['label']} 承接 {format_percent(rating_barrier.get('share'))} 销量；系统补充样本评分数 {format_number(supplement.get('ratings'))}。"),
            ("T 威胁", f"{ip_terms} 仅取得风险信号；命名、兼容表达和结构方案必须在立项前完成法务复核。"),
        ]
    fills = {"S": LIGHT_GREEN, "W": LIGHT_RED, "O": PALE_BLUE, "T": YELLOW}
    for row, (label, text) in enumerate(rows, 4):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, text)
        style_range(sheet, f"A{row}:B{row}", fills[label[0]], TEXT, label[0] in "SWOT")
        sheet.row_dimensions[row].height = 42
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 115
    sheet.freeze_panes = "A4"
    setup_page(sheet, "A1:B11")


def render_aba(sheet, data: dict) -> None:
    reset_sheet(sheet)
    sheet.sheet_view.showGridLines = True
    sheet.print_area = None


def render_raw(sheet, data: dict, validation: dict | None = None) -> None:
    reset_sheet(sheet)
    merge_title(sheet, "A1:M1", "MCP 原始数据摘要｜已脱敏、聚合、无原始评论文本")
    merge_note(sheet, "A2:M2", "本 Sheet 保留可审计字段、口径和证据 ID；不包含 API key、密钥 URL、绝对路径或评论原文。", LIGHT_BLUE)
    merge_note(sheet, "A4:M4", "市场快照", BLUE, 11)
    header(sheet, 5, 1, 12, ["月份", "节点", "产品数", "品牌数", "卖家数", "均价", "月均销量", "月均销售额", "平均BSR", "平均评分数", "平均评分", "新品占比"])
    snap = data["market"]["snapshot"]
    values = [snap["month"], data["queryManifest"]["selectedNodeIdPath"], snap["products"], snap["brands"], snap["sellers"], snap["avgPrice"], snap["avgUnits"], snap["avgRevenue"], snap["avgBsr"], snap["avgRatings"], snap["avgRating"], snap["newProductShare"]]
    for col, value in enumerate(values, 1):
        sheet.cell(6, col, value)
    sheet.cell(6, 12).number_format = "0.00%"

    merge_note(sheet, "A8:M8", "12 个月类目统计", BLUE, 11)
    header(sheet, 9, 1, 8, ["月份", "月均销量", "月均销售额", "平均BSR", "均价", "新品占比", "新品均销", "证据ID"])
    for row, item in enumerate(data["market"]["monthlyTrend"], 10):
        values = [item["month"], item["avgUnits"], item["avgRevenue"], item["avgBsr"], item["avgPrice"], item["newProductShare"], item["newAvgUnits"], item["evidenceId"]]
        for col, value in enumerate(values, 1):
            sheet.cell(row, col, value)
        sheet.cell(row, 6).number_format = "0.00%"
    alternating(sheet, 10, 21, 1, 8)

    merge_note(sheet, "A23:M23", "四款代表竞品", BLUE, 11)
    header(sheet, 24, 1, 13, ["ASIN", "来源", "品牌", "售价", "评分", "评分数", "评论数", "履约", "上架时间", "节点", "角色", "销量快照", "证据ID"])
    for row, item in enumerate(data["competitors"], 25):
        numeric_values = [number_or_none(item.get(key)) for key in ("price", "rating", "ratings", "reviews", "salesSnapshot")]
        numeric_values = [value if value is not None else "N/A" for value in numeric_values]
        evidence_ids = item.get("evidenceIds") or []
        values = [
            text_or_na(item.get("asin")), text_or_na(item.get("source")), text_or_na(item.get("brand")),
            numeric_values[0], numeric_values[1], numeric_values[2], numeric_values[3], text_or_na(item.get("fulfillment")),
            text_or_na(item.get("listingDate")), text_or_na(item.get("nodeIdPath")), text_or_na(item.get("role")),
            numeric_values[4], "; ".join(str(value) for value in evidence_ids) or "N/A",
        ]
        for col, value in enumerate(values, 1):
            sheet.cell(row, col, value)
    alternating(sheet, 25, 28, 1, 13)

    merge_note(sheet, "A30:M30", "关键词摘要", BLUE, 11)
    header(sheet, 31, 1, 9, ["关键词", "搜索量", "购买量", "购买率", "PPC", "商品数", "供需比", "工具", "证据ID"])
    for row, item in enumerate(data["keywords"]["items"], 32):
        tool = item.get("tool", "keyword_miner")
        values = [
            text_or_na(item.get("keyword")),
            *[(value if value is not None else "N/A") for value in (number_or_none(item.get(key)) for key in ("searches", "purchases", "purchaseRate", "ppc", "products", "supplyDemandRatio"))],
            text_or_na(tool), text_or_na(item.get("evidenceId")),
        ]
        for col, value in enumerate(values, 1):
            sheet.cell(row, col, value)
        sheet.cell(row, 4).number_format = "0.00%"
    alternating(sheet, 32, 38, 1, 9)

    merge_note(sheet, "A40:M40", "数据缺口", BLUE, 11)
    header(sheet, 41, 1, 4, ["字段", "缺口", "影响", "处理"])
    for row, item in enumerate(data["dataGaps"], 42):
        for col, key in enumerate(("field", "description", "impact", "handling"), 1):
            sheet.cell(row, col, item[key])
        sheet.row_dimensions[row].height = 48
    alternating(sheet, 42, 41 + len(data["dataGaps"]), 1, 4)
    end_row = 41 + len(data["dataGaps"])
    if validation and validation.get("consensus"):
        summary_row = end_row + 2
        merge_note(sheet, f"A{summary_row}:M{summary_row}", "Gemini + GLM 脱敏交叉验证", BLUE, 11)
        header(sheet, summary_row + 1, 1, 4, ["状态", "共识", "平均分", "模型结果"])
        consensus = validation["consensus"]
        models = "；".join(
            f"{item['provider']} / {item['model']}: {item['verdict']} {item['score_0_to_5']}/5"
            for item in consensus.get("modelVerdicts", [])
        )
        sheet.cell(summary_row + 2, 1, validation.get("status", "N/A"))
        sheet.cell(summary_row + 2, 2, consensus.get("verdict", "N/A"))
        sheet.cell(summary_row + 2, 3, consensus.get("score_0_to_5", "N/A"))
        sheet.cell(summary_row + 2, 4, models)
        style_range(sheet, f"A{summary_row + 2}:D{summary_row + 2}", YELLOW)
        end_row = summary_row + 2
    sheet.column_dimensions["A"].width = 30
    for col in range(2, 14):
        sheet.column_dimensions[get_column_letter(col)].width = 20
    sheet.freeze_panes = "A5"
    setup_page(sheet, f"A1:M{end_row}")


def render_sources(sheet, data: dict) -> None:
    reset_sheet(sheet)
    merge_title(sheet, "A1:E1", "数据来源与采用状态")
    header(sheet, 3, 1, 5, ["证据ID", "MCP工具", "查询范围", "采用状态", "说明"])
    for row, item in enumerate(data["dataSources"], 4):
        note = item.get("note") or (
            "接口未采用或空返回；仅登记数据缺口，不推断"
            if str(item.get("status", "")).lower() in {"empty", "not_used", "failed", "缺失", "未采用"}
            else "用于报告与 Excel 的对应事实字段"
        )
        values = [item["evidenceId"], item["tool"], item["scope"], item["status"], note]
        for col, value in enumerate(values, 1):
            sheet.cell(row, col, value)
        sheet.row_dimensions[row].height = 34
        style_range(sheet, f"A{row}:E{row}", LIGHT_GRAY if row % 2 == 0 else WHITE)
    widths = [20, 38, 52, 24, 48]
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A4"
    setup_page(sheet, f"A1:E{3 + len(data['dataSources'])}")


def render_markdown(data: dict, qa: dict | None, validation: dict | None = None) -> str:
    title = data["render"]["markdownTitle"]
    product = data["productDirection"]
    selected = data["categorySelection"]["selected"]
    snap = data["market"]["snapshot"]
    decision = data["decision"]
    qastatus = qa.get("status", "NOT_RUN") if qa else "NOT_RUN"
    validation_status = validation.get("status", "NOT_RUN") if validation else "NOT_RUN"
    consensus = validation.get("consensus", {}) if validation else {}
    trend = data["market"]["monthlyTrend"]
    units = [number_or_zero(item.get("avgUnits")) for item in trend]
    avg_units = sum(units) / len(units) if units else 0
    peak = max(trend, key=lambda item: number_or_zero(item.get("avgUnits")), default={"month": "N/A", "avgUnits": 0})
    price_bucket = max(data["market"]["distributions"]["price"]["items"], key=lambda item: number_or_zero(item.get("share")), default={"label": "N/A", "share": 0})
    ratings_bucket = max(data["market"]["distributions"]["ratingCount"]["items"], key=lambda item: number_or_zero(item.get("share")), default={"label": "N/A", "share": 0})
    rating_bucket = max(data["market"]["distributions"]["rating"]["items"], key=lambda item: number_or_zero(item.get("share")), default={"label": "N/A", "share": 0})
    keyword_items = data["keywords"]["items"]
    top_keyword = max(keyword_items, key=lambda item: number_or_zero(item.get("searches")), default={"keyword": "N/A", "searches": 0})
    top_purchase = max(keyword_items, key=lambda item: number_or_zero(item.get("purchaseRate")), default={"keyword": "N/A", "purchaseRate": 0})
    lowest_supply = min(keyword_items, key=lambda item: number_or_zero(item.get("supplyDemandRatio")), default={"keyword": "N/A", "supplyDemandRatio": 0})
    model_results = "；".join(
        f"{item.get('provider')} / {item.get('model')}: {item.get('verdict')} {item.get('score_0_to_5')}/5"
        for item in consensus.get("modelVerdicts", [])
    ) or "N/A"
    model_issues: list[str] = []
    if validation:
        for judge in validation.get("judges", []):
            for issue in (judge.get("parsed") or {}).get("critical_issues", []):
                if issue not in model_issues:
                    model_issues.append(issue)
    lines = [
        f"# {title}", "", f"> 站点：{data['command']['site']}  ",
        f"> 种子竞品：{'、'.join(data['command']['seedAsins'])}  ",
        f"> 系统补充竞品：{data['command']['systemSupplement']}  ",
        f"> 数据月份：{data['queryManifest']['lastCompleteMonth']}", "",
        "## 立项结论", "",
        f"**状态：{decision['status']}。** {decision['summary']}", "",
        "## 产品意向", "",
        f"- 产品方向：{product['description']}。",
        f"- 核心用户：{product['targetAudience']}。",
        f"- 小样测试价格：{product['targetPrice']}；这是验证价格带，不是利润结论。", "",
        "已验证功能：", "",
    ]
    lines.extend(f"- {item}" for item in product["verifiedFunctions"])
    lines.extend(["", "待验证未满足需求：", ""])
    lines.extend(f"- {item}" for item in product["unmetNeeds"])
    lines.extend([
        "", "## 类目与市场", "",
        f"主分析节点：`{selected['labelPath']}`（`{selected['nodeIdPath']}`）。",
        "", f"选择依据：{text_or_na(selected.get('reason'))}。{text_or_na(selected.get('overrideReason'))}。", "",
        f"{text_or_na(snap.get('month'))} 快照：{format_number(snap.get('products'))} 个样本、{format_number(snap.get('brands'))} 个品牌、{format_number(snap.get('sellers'))} 个卖家、均价 `{format_money(snap.get('avgPrice'))}`、平均月销 `{format_number(snap.get('avgUnits'))}`、平均销售额 `{format_money(snap.get('avgRevenue'))}`、平均评分数 `{format_number(snap.get('avgRatings'))}`、平均评分 `{format_number(snap.get('avgRating'), 1)}`、新品占比 `{format_percent(snap.get('newProductShare'), 0)}`。",
        "", f"- 连续 {len(trend)} 个月平均月销约 {avg_units:,.0f} 件，峰值为 {peak['month']} 的 {format_number(peak.get('avgUnits'))} 件。",
        f"- 销量占比最高价格带为 `{price_bucket['label']}`（{format_percent(price_bucket.get('share'))}）；测试价格 `{product['targetPrice']}` 必须由可验证差异支撑。",
        f"- 评分数区间 `{ratings_bucket['label']}` 承接 {format_percent(ratings_bucket.get('share'))} 销量，评分值区间 `{rating_bucket['label']}` 承接 {format_percent(rating_bucket.get('share'))} 销量。",
        f"- 品牌 CR5 月销量占比 {format_percent(data['market'].get('concentration', {}).get('brandCr5Units'))}，需要结合品牌分布判断集中度。",
        "", "## 代表竞品与 VOC", "",
        "| ASIN | 角色 | 售价 | 评分/评分数 | 已验证卖点 | 低星风险 |",
        "|---|---:|---:|---:|---|---|",
    ])
    for item in data["competitors"]:
        lines.append(
            f"| {text_or_na(item.get('asin'))} | {text_or_na(item.get('role'))} | {format_money(item.get('price'))} | "
            f"{format_number(item.get('rating'), 1)}/{format_number(item.get('ratings'))} | {text_or_na(item.get('features'))} | "
            f"{text_or_na(item.get('negativeThemes'))} |"
        )
    lines.extend(["", "## 关键词与流量证据", "", "> `ABA排名【季度】` Sheet 按标准模板保留为空白，由业务人员人工维护。", "", f"| 关键词 | {data['queryManifest']['lastCompleteMonth']}搜索量 | 购买量 | 购买率 | PPC | 商品数 |", "|---|---:|---:|---:|---:|---:|"])
    for item in data["keywords"]["items"]:
        lines.append(
            f"| {text_or_na(item.get('keyword'))} | {format_number(item.get('searches'))} | {format_number(item.get('purchases'))} | "
            f"{format_percent(item.get('purchaseRate'))} | {format_money(item.get('ppc'))} | {format_number(item.get('products'))} |"
        )
    missing_text = "、".join(data["keywords"].get("missing", [])) or "无"
    lines.extend([
        "",
        f"- 搜索量最高词：`{top_keyword['keyword']}`（{format_number(top_keyword.get('searches'))}）。",
        f"- 购买率最高词：`{top_purchase['keyword']}`（{format_percent(top_purchase.get('purchaseRate'))}）。",
        f"- 供需比最低词：`{lowest_supply['keyword']}`（{format_number(lowest_supply.get('supplyDemandRatio'), 2)}），需要谨慎评估竞争范围。",
        f"- 未返回关键词：{missing_text}；不补值。", "", "## 决策门禁", "",
    ])
    for gate in decision["gates"]:
        evidence_ids = gate.get("evidenceIds") or []
        lines.append(f"- **{text_or_na(gate.get('name'))} / {text_or_na(gate.get('status'))}**：阈值：{text_or_na(gate.get('threshold'))}；实际：{text_or_na(gate.get('actual'))}；证据：`{'; '.join(str(value) for value in evidence_ids) or 'N/A'}`。")
    lines.extend(["", "## P0-P2 动作", ""])
    for item in decision["actions"]:
        evidence_ids = item.get("evidenceIds") or []
        lines.append(f"- **{text_or_na(item.get('priority'))}**：{text_or_na(item.get('action'))}（证据：`{'; '.join(str(value) for value in evidence_ids) or 'N/A'}`）")
    lines.extend(["", "## 数据缺口", ""])
    for item in data["dataGaps"]:
        lines.append(f"- **{item['field']}**：{item['description']}；影响：{item['impact']}；处理：{item['handling']}。")
    lines.extend([
        "", "## 验证摘要", "",
        f"- Python 统一证据规则校验：`{qastatus}`。",
        f"- Gemini + GLM 脱敏交叉验证：`{validation_status}`；共识 `{consensus.get('verdict', 'N/A')}`，平均分 `{consensus.get('score_0_to_5', 'N/A')} / 5`。",
        f"- 模型结果：{model_results}。",
        f"- 模型重点问题：{'；'.join(model_issues[:3]) if model_issues else 'N/A'}。",
        "- 验证输入禁止包含 API key、token、MCP 密钥 URL、绝对路径、原始评论和私有供应链数据。",
        "", "本企划用于小样立项讨论。工程实现、供应链、成本、物流、专利和合规需在量产决策前单独验证。", "",
    ])
    return "\n".join(lines)


def verify_outputs(workbook_path: Path, markdown_path: Path, data: dict, image_required: bool) -> dict:
    workbook = load_workbook(workbook_path, data_only=False)
    errors: list[str] = []
    warnings: list[str] = []
    if workbook.sheetnames != SHEET_ORDER:
        errors.append(f"Sheet 顺序错误: {workbook.sheetnames}")
    if len(workbook["市场分析"]._charts) != 7:
        errors.append("市场分析图表数量不是 7")
    if len(workbook["竞品分析与优化策略"]._charts) != 1:
        errors.append("竞品双轴图缺失")
    aba_sheet = workbook["ABA排名【季度】"]
    if any(cell.value is not None for row in aba_sheet.iter_rows() for cell in row):
        errors.append("ABA 人工维护 Sheet 不是空白")
    if aba_sheet._charts or aba_sheet._images or aba_sheet.merged_cells.ranges:
        errors.append("ABA 人工维护 Sheet 存在残留图表、图片或合并区域")
    if image_required and len(workbook["意向产品"]._images) != 1:
        warnings.append("首个 ASIN 主图未嵌入")
    values = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    values.append(cell.value)
    text = "\n".join(values) + "\n" + markdown_path.read_text(encoding="utf-8")
    expected_node = str(data["queryManifest"]["selectedNodeIdPath"])
    if expected_node not in str(workbook["市场分析"]["A1"].value):
        errors.append("市场分析未使用 evidence.json 中锁定的类目节点")
    expected_supplement = str(data["command"]["systemSupplement"])
    competitor_asins = {str(workbook["竞品分析与优化策略"].cell(row, 1).value) for row in range(6, 10)}
    if expected_supplement not in competitor_asins:
        errors.append("竞品表未使用 evidence.json 中的系统补充样本")
    if SECRET_RE.search(text):
        errors.append("检测到疑似密钥")
    if ABS_PATH_RE.search(text):
        errors.append("检测到绝对本地路径")
    if PLACEHOLDER_RE.search(text):
        errors.append("检测到模板占位符")
    with zipfile.ZipFile(workbook_path) as archive:
        for member in archive.namelist():
            if member.endswith(".xml"):
                ElementTree.fromstring(archive.read(member))
    return {"status": "FAIL" if errors else ("WARN" if warnings else "PASS"), "errors": errors, "warnings": warnings}


def main() -> int:
    args = parse_args()
    data = load_json(args.evidence)
    qa = load_json(args.qa) if args.qa and args.qa.exists() else None
    validation = load_json(args.validation) if args.validation and args.validation.exists() else None
    args.output_dir.mkdir(parents=True, exist_ok=True)
    title = data["render"]["workbookTitle"]
    workbook_path = args.output_dir / f"{title}.xlsx"
    markdown_path = args.output_dir / f"{title}.md"
    preserved_image = None
    if workbook_path.exists():
        existing = load_workbook(workbook_path)
        if existing["意向产品"]._images:
            preserved_image = existing["意向产品"]._images[0]._data()
    shutil.copy2(args.template, workbook_path)
    workbook = load_workbook(workbook_path)
    if workbook.sheetnames != SHEET_ORDER:
        raise ValueError(f"模板 Sheet 顺序不符: {workbook.sheetnames}")
    image_embedded = render_intent(workbook["意向产品"], data, preserved_image)
    render_market(workbook["市场分析"], data)
    old_competitor = workbook["竞品分析与优化策略"]
    competitor_index = workbook.index(old_competitor)
    workbook.remove(old_competitor)
    competitor_sheet = workbook.create_sheet("竞品分析与优化策略", competitor_index)
    render_competitors(competitor_sheet, data)
    render_swot(workbook["SWOT分析"], data)
    old_aba = workbook["ABA排名【季度】"]
    aba_index = workbook.index(old_aba)
    workbook.remove(old_aba)
    aba_sheet = workbook.create_sheet("ABA排名【季度】", aba_index)
    render_aba(aba_sheet, data)
    render_raw(workbook["MCP原始数据"], data, validation)
    render_sources(workbook["数据来源"], data)
    workbook.save(workbook_path)
    markdown_path.write_text(render_markdown(data, qa, validation), encoding="utf-8")
    result = verify_outputs(workbook_path, markdown_path, data, image_required=True)
    result["imageEmbedded"] = image_embedded
    (args.output_dir / "render_qa.json").write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({"workbook": workbook_path.name, "markdown": markdown_path.name, "qa": result}, ensure_ascii=False, indent=2))
    return 1 if result["status"] == "FAIL" else 0


if __name__ == "__main__":
    raise SystemExit(main())
